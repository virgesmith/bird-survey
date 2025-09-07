import json
from pathlib import Path
from typing import Any

import openpyxl
from itrx import Itr
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from model import CLOUD_KEY, RAIN_KEY, VISIBILITY_KEY, WIND_KEY, SurveyData, Surveys


def safe_get(choices: list[str], index: int) -> str | int:
    if 0 <= index < len(choices):
        return choices[index]
    return index


def apply_border(ws, start_col: int, end_col: int, start_row: int, end_row: int) -> None:
    border = Side(border_style="thin", color="000000")

    cell = ws.cell(row=start_row, column=start_col)
    cell.border = Border(top=border, left=border)

    cell = ws.cell(row=end_row, column=start_col)
    cell.border = Border(bottom=border, left=border)

    cell = ws.cell(row=start_row, column=end_col)
    cell.border = Border(top=border, right=border)

    cell = ws.cell(row=end_row, column=end_col)
    cell.border = Border(bottom=border, right=border)

    for c in range(start_col + 1, end_col):
        cell = ws.cell(row=start_row, column=c)
        cell.border = Border(top=border)

        cell = ws.cell(row=end_row, column=c)
        cell.border = Border(bottom=border)

    for r in range(start_row + 1, end_row):
        cell = ws.cell(row=r, column=start_col)
        cell.border = Border(left=border)

        cell = ws.cell(row=r, column=end_col)
        cell.border = Border(right=border)


def apply_bold(ws, start_col: int, end_col: int, start_row: int, end_row: int):
    bold = Font(name="Calibri", bold=True)
    for c in range(start_col, end_col + 1):
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=c).font = bold


def export_to_excel_sheet(ws: Any, surveys: Surveys | tuple[SurveyData, ...]) -> None:
    # Header rows
    ws["B1"] = f"Transect number: {ws.title}"
    ws["B1"].font = Font(name=ws["B1"].font.name, bold=True)
    ws["B2"] = "km square grid reference:"

    # Segment headers
    segment_count = 12
    col = 3
    side = "left"
    for seg in range(segment_count * 2):
        ws.cell(row=2, column=col, value=f"segment {seg // 2 + 1} - {side}")
        ws.cell(row=3, column=col, value="species")
        ws.cell(row=3, column=col + 1, value="count")
        col += 2
        side = "left" if side == "right" else "right"
    apply_border(ws, 3, 2 + 4 * segment_count, 2, 3)

    # Fill visit rows
    base_row = 4
    for visit_idx, survey in enumerate(surveys):
        row = base_row + visit_idx * 14
        ws.cell(row=row, column=1, value=f"Visit {visit_idx + 1}")
        ws.cell(row=row + 1, column=1, value="Date:")
        ws.cell(row=row + 2, column=1, value="surveyors:")
        ws.cell(row=row + 4, column=1, value="start time:")
        ws.cell(row=row + 5, column=1, value="end time:")
        ws.cell(row=row + 6, column=1, value="start time:")
        ws.cell(row=row + 7, column=1, value="end time:")
        ws.cell(row=row + 8, column=1, value="Weather")
        ws.cell(row=row + 9, column=1, value="cloud:")
        ws.cell(row=row + 10, column=1, value="rain:")
        ws.cell(row=row + 11, column=1, value="wind:")
        ws.cell(row=row + 12, column=1, value="visibility:")

        # Fill survey metadata
        ws.cell(row=row + 1, column=2, value=survey.visit_date)
        ws.cell(row=row + 2, column=2, value=survey.observer_name)
        ws.cell(row=row + 4, column=2, value=survey.first_segment_start_time)
        ws.cell(row=row + 5, column=2, value=survey.first_segment_end_time)
        ws.cell(row=row + 6, column=2, value=survey.second_segment_start_time)
        ws.cell(row=row + 7, column=2, value=survey.second_segment_end_time)

        ws.cell(row=row + 9, column=2, value=safe_get(CLOUD_KEY, survey.weather_code.cloud))
        ws.cell(row=row + 10, column=2, value=safe_get(RAIN_KEY, survey.weather_code.rain))
        ws.cell(row=row + 11, column=2, value=safe_get(WIND_KEY, survey.weather_code.wind))
        ws.cell(row=row + 12, column=2, value=safe_get(VISIBILITY_KEY, survey.weather_code.visibility))

        ws.cell(row=row + 1, column=2).font = Font(bold=True)
        apply_bold(ws, 2, 2, row + 1, row + 12)
        apply_border(ws, 1, 2, row, row + 13)
        apply_border(ws, 3, 2 + segment_count * 4, row, row + 13)

        # Fill sightings for each segment
        for segment in survey.segments:
            seg_idx = segment.number
            col_base = -1 + seg_idx * 4
            ws.cell(row=2, column=col_base + 1, value=segment.start_coordinate)
            ws.cell(row=2, column=col_base + 3, value=segment.end_coordinate)
            # Left
            for sight_idx, sight in enumerate(segment.left):
                ws.cell(row=row + sight_idx, column=col_base, value=sight.code.bird_name)
                ws.cell(row=row + sight_idx, column=col_base + 1, value=sight.count or 1)
            # Right
            for sight_idx, sight in enumerate(segment.right):
                ws.cell(row=row + sight_idx, column=col_base + 2, value=sight.code.bird_name)
                ws.cell(row=row + sight_idx, column=col_base + 3, value=sight.count or 1)

    # format columns and fill
    # Iterate through all columns in the worksheet, use the column_dimensions property to get the column object,
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15
    ws.column_dimensions["B"].width = 28

    alternate = False
    for col in ws.iter_cols(min_row=3, max_row=3 + len(surveys) * 14, min_col=3, max_col=2 + segment_count * 4):
        bg = "BBBBBB" if alternate else "DDDDDD"
        for cell in col:
            cell.alignment = Alignment(horizontal="left")
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        alternate = not alternate


def export_to_excel(all_surveys: Surveys) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # TODO also sort by date
    for transect, surveys in Itr(all_surveys).groupby(lambda s: s.transect_number).collect(dict).items():
        print(f"Transect {transect} {len(surveys)} surveys")
        surveys = sorted(surveys, key=lambda s: s.visit_date)
        ws = wb.create_sheet(title=f"Transect {transect}")
        export_to_excel_sheet(ws, surveys)

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    return wb


def main() -> None:
    surveys = Surveys([])
    for file in Path("./data").glob("*.json"):
        with open(file) as fd:
            surveys.extend(Surveys(json.load(fd)))
    wb = export_to_excel(surveys)
    wb.save("./data/all_surveys.xls")


if __name__ == "__main__":
    main()
